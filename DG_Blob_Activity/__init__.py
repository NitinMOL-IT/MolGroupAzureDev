# This function is not intended to be invoked directly. Instead it will be
# triggered by an orchestrator function.
# Before running this sample, please:
# - create a Durable orchestration function
# - create a Durable HTTP starter function
# - add azure-functions-durable to requirements.txt
# - run pip install -r requirements.txt

import azure.functions as func
import logging
import json
from .az_datalake_activity import DataLakeUtility
import os
# import datetime
from datetime import datetime, timedelta
# import datetime
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils import get_column_letter


def main(name: str) -> str:
        obj_DataLakeUtility = DataLakeUtility()

        today_date = datetime.now().strftime('%Y-%m-%d')
        blob_path = f"DGBoiler_Log_Folder/{today_date}"
        destination_container_name = "nonprdfocusbisto"

        log_list = []

        log_list.append("Variable initialization has started")

        Blob_Trigger_Initialization_Log = obj_DataLakeUtility.upload_log(log_data_string=str(log_list), blob_name="Blob_Trigger_Initialization_Log.log",blob_path=blob_path,destination_container_name=destination_container_name)

        try:
                log_Activity = []
                log_Activity.append("Blob trigger function has been started.")
                # logging.info(f"Python blob trigger function processed blob" f"Name: {myblob.name}" f"Blob Size: {myblob.length} bytes")

                #-------------------Missing Vessel Data Archiving Start-----------------------------
                # log_Activity.append("Missing Vessel Data Archiving Start")

                # obj_DataLakeUtility.missing_report_archieve()

                # log_Activity.append("Missing Vessel Data Archiving End")
                #-------------------Missing Vessel Data Archiving End-----------------------------
                log_Activity.append("Downloading Files to temp Folder")

                Event_Consumption_file_download = obj_DataLakeUtility.download_blob("molpearlqrt",
                                                                                "Online_ABLOG/DG-Boiler/Event&OliConsumption.csv", "Event&OliConsumption.csv")
                
                Active_Voyage_file_download = obj_DataLakeUtility.download_blob("molpearlqrt", "Online_ABLOG/DG-Boiler/ActiveVoyages.csv", "ActiveVoyages.csv")

                Complete_Vessel_file_download = obj_DataLakeUtility.download_blob("molpearlqrt","Online_ABLOG/DG-Boiler/VesselData.csv", "VesselData.csv")

                Final_processed_file_download = obj_DataLakeUtility.download_blob("nonprdfocusbisto", "ABLOG_FINAL/DG-Boiler/DGBoiler_processed_data.csv", "DGBoiler_processed_data.csv")

                Final_Missing_file_download = obj_DataLakeUtility.download_blob("nonprdfocusbisto","ABLOG_FINAL/DG-Boiler/Missing DG-Boiler Reports Unpivoted.csv", "Missing DG-Boiler Reports Unpivoted.csv")

                log_Activity.append("Downloading completed Files to temp Folder")

                # #File paths
                Event_consumption_file_path = Event_Consumption_file_download['absolute_path']
                Active_Voyage_file_path = Active_Voyage_file_download['absolute_path']
                Complete_vessel_file_path = Complete_Vessel_file_download['absolute_path']
                Final_processed_file_path = Final_processed_file_download['absolute_path']
                Final_Missing_file_path = Final_Missing_file_download['absolute_path']

                #Reading the files
                log_Activity.append("Reading the Files from temp Folder")

                df_EC = pd.read_csv(Event_consumption_file_path, header=0)
                log_Activity.append("Event&Consumption file read complete")

                df_AV = pd.read_csv(Active_Voyage_file_path, header=0)
                log_Activity.append("ActiveVoyage file read complete")

                df_vsl = pd.read_csv(Complete_vessel_file_path, header=0)
                log_Activity.append("Complete Vessel List file read complete")

                df_missing = pd.read_csv(Final_Missing_file_path, header=0)
                log_Activity.append("Missing DG-Boiler file read complete")
                
                #-------------------------------------Commented Old Code--------------------------------------------
                # df_EC2 = df_EC[['VoyageID', 'VoyageNo','VesselID','VesselName']] #create new dataframe with Voyage and Vessel info
                # log_Activity.append("created new dataframe with Voyage and Vessel info")

                # # List of columns to consider for identifying duplicates
                # columns_to_check_duplicates = ["VoyageID", "VoyageNo", "VesselID","VesselName"]

                # # Remove duplicates based on the specified columns
                # df_EC2 = df_EC2.drop_duplicates(subset=columns_to_check_duplicates, keep="first").reset_index(drop=True)
                # log_Activity.append("Removed duplicates based on the specified columns")

                # #left join with Active Voyages
                # df_all = df_AV.merge(df_EC2, on=['VoyageID'], how='left', indicator=True)
                # log_Activity.append("Left join complete")

                # #create Dataframe with rows that exist in Active Voyage only (Missing data identification)

                # df_missing = df_all[df_all['_merge']=='left_only']
                # log_Activity.append("Missing Noon Report Vessels are identified")

                # df_missing2 = df_missing[['VoyageID', 'VoyageNo_x', 'Status', 'Vessel Name']]
                # today = datetime.today().date()
                
                # df_missing2['Date'] = today - timedelta(days=1)
                #---------------------------------------Commented Old Code-------------------------------

                df_EC2 = df_EC.groupby(["IMONo", "VesselName","Event Date"], as_index=False).agg({"MachineID": pd.Series.nunique}) #grouping based on IMo, Vessel Name and Event Date to understand how many machines are used
                log_Activity.append('grouping based on IMo, Vessel Name and Event Date to understand how many machines are used')

                try:
                        df_EC2['Event Date'] = pd.to_datetime(df_EC2['Event Date'], format='%d-%m-%Y').dt.strftime('%d-%m-%Y')
                except:
                        df_EC2['Event Date'] = pd.to_datetime(df_EC2['Event Date'], format='%Y-%m-%d').dt.strftime('%d-%m-%Y')

                df_all = pd.merge(df_vsl,df_EC2, on = 'IMONo', how='left', suffixes=('_vsl','_EO')) # Joining with Targetted Vessels to find out missing vessels
                log_Activity.append('Joining with Targetted Vessels to find out missing vessels done')

                df_all['Report Status'] = ['ok' if m>0 else 'missing' for m in df_all['MachineID']] # identifying the vessels with 'ok' and 'missing'
                df_all.rename(columns={'VesselName_vsl': 'Vessel Name'}, inplace=True) #Renaming the Vessel Name column
                log_Activity.append('Identifying the missing vessels and column renaming done')

                #---------------------------------------------------------------------------------------------------------------------------
                #                                              Storing the data in Unpivoted Format Start                                   |
                #---------------------------------------------------------------------------------------------------------------------------
                df_missing2 = pd.concat([df_missing,df_all], ignore_index=True) # Merging today's missing data with previous data
                log_Activity.append('Merging missing data of today with previous data done for Unpivoted Data')

                temp_missing_data_file_loc = os.path.join(obj_DataLakeUtility.temp_directory,'Missing_Vessel_Info.csv')
                df_missing2.to_csv(temp_missing_data_file_loc, index=False)
                log_Activity.append('Missing Vessel Data has been converted to CSV')

                Missing_data_file_upload = obj_DataLakeUtility.upload_blob(destination_container_name, blob_path="ABLOG_FINAL/DG-Boiler", temp_file_abs_path=temp_missing_data_file_loc, blob_name="Missing DG-Boiler Reports Unpivoted.csv")

                log_Activity.append("Missing Vessel data has been uploaded in Blob location")

                #---------------------------------------------------------------------------------------------------------------------------
                #                                           Storing the data in Unpivoted Format End                                       |
                #---------------------------------------------------------------------------------------------------------------------------

                #---------------------------------------------------------------------------------------------------------------------------
                #                                          Start Pivoting the data for preparing the Excel Format Report                  |
                #---------------------------------------------------------------------------------------------------------------------------

                df_pivot = df_missing2.pivot_table(index=['Division','IMONo','Vessel Name'], columns='Event Date', values='Report Status', aggfunc=np.min) #Pivoting the data based Event Date
                log_Activity.append('Pivoting the data based Event Date done')

                for col in df_pivot:
                        df_pivot[col] = df_pivot[col].astype(str)
                        df_pivot[col].replace('nan','missing', inplace=True)

                temp_missing_excel_data_file_loc = os.path.join(obj_DataLakeUtility.temp_directory,'Final Missing Vessel Pivoted.xlsx')

                df_pivot.to_excel(temp_missing_excel_data_file_loc, index=True)
                
                log_Activity.append('Final missing vessel data in pivoted format is exported to Excel')

                wb = openpyxl.load_workbook(temp_missing_excel_data_file_loc) #Read the excel file
                ws = wb.active #activating the first worksheet

                #Specify the range of columns to check (D to ZZ)
                start_column_index = 4  # Column D
                # end_column_index = 702  # Column ZZ
                columns_to_check = [get_column_letter(i) for i in range(start_column_index, ws.max_column + 1)]

                log_Activity.append("Got the range of rows and columns in the worksheet")

                # Iterate through the specified columns
                for column_letter in columns_to_check:
                      current_column = ws[column_letter]
                      for cell in current_column:                             
                             if cell.value == "missing":                                     
                                     cell.font = Font(color="FF0000")  # Set font color to red


                log_Activity.append('Coloring the cells where the value is mentioned as missing')

                # temp_pivot_missing_data_file_loc = os.path.join(obj_DataLakeUtility.temp_directory,'Missing DG-Boiler Reports_Pivot.xlsx')
                wb.save(temp_missing_excel_data_file_loc)
                log_Activity.append('Pivoted Missing Vessel Data has been converted to Excel')

                Missing_data_file_upload = obj_DataLakeUtility.upload_blob(destination_container_name, blob_path="ABLOG_FINAL/DG-Boiler", temp_file_abs_path=temp_missing_excel_data_file_loc, blob_name="Missing DG-Boiler Excel Reports.xlsx")

                #---------------------------------------------------------------------------------------------------------------------------
                #                                          End Pivoting the data for preparing the Excel Format Report                  |
                #---------------------------------------------------------------------------------------------------------------------------

                # Define a dictionary for column data types
                column_data_types = {
                        "EventID": 'str',                        
                        "SpanTime2": 'str',
                        "VoyageID": 'int64',
                        "VoyageNo" : 'str',
                        "VesselID": 'int64',
                        "VesselName": 'str',
                        "IsWindDirectionNA": 'str',
                        "IsSwellDirectionNA" : 'str',
                        "SeawaterTemperature" : 'float64',
                        "EventName": 'str',
                        "SpanName" : 'str',
                        "IsDG1Used": 'str',
                        "IsDG2Used": 'str',
                        "IsDG3Used": 'str',
                        "IsDG4Used": 'str',
                        "IsDG5Used": 'str',
                        "IsDG6Used": 'str',
                        "AveOutput": 'float64',
                        "oil_name": 'str',
                        "MachineID": 'int64',
                        "machine_name": 'str',
                        "OilConsumptionRate": 'float64',
                        "IMONo": 'str',
                        "Division":'str'
                }
                # Apply the data type conversion to the DataFrame
                df_EC = df_EC.astype(column_data_types)
                df_EC['EventDateTime'] = pd.to_datetime(df_EC['EventDateTime'], utc=True, format='%Y-%m-%d %H:%M:%S.%f %z')
                df_EC['SpanTime'] = pd.to_datetime(df_EC['SpanTime'], utc=True, format='%Y-%m-%d %H:%M:%S.%f %z')
                try:
                        df_EC['Event Date'] = pd.to_datetime(df_EC['Event Date'], format='%d-%m-%Y').dt.strftime('%d-%m-%Y')
                except:
                        df_EC['Event Date'] = pd.to_datetime(df_EC['Event Date'], format='%Y-%m-%d').dt.strftime('%d-%m-%Y')

                df_EC['SpanTime3'] = df_EC['SpanTime2'].str.split(' ', expand=True)
                log_Activity.append("Changed Datatypes of columns")

                df_EC['IsDG1Used'] = df_EC['IsDG1Used'].replace('nan', 'Null')
                df_EC['IsDG2Used'] = df_EC['IsDG2Used'].replace('nan', 'Null')
                df_EC['IsDG3Used'] = df_EC['IsDG3Used'].replace('nan', 'Null')
                df_EC['IsDG4Used'] = df_EC['IsDG4Used'].replace('nan', 'Null')
                df_EC['IsDG5Used'] = df_EC['IsDG5Used'].replace('nan', 'Null')
                df_EC['IsDG6Used'] = df_EC['IsDG6Used'].replace('nan', 'Null')

                DG_Usage_value_mapping = {
                        "TRUE" : "1",
                        "True": "1",
                        "False": "0",
                        "FALSE" : "0",
                        "Null" : "0"
                }

                def replace_value(value):
                        return DG_Usage_value_mapping.get(value, value)
                
                df_EC['IsDG1Used'] = df_EC['IsDG1Used'].apply(replace_value)
                df_EC['IsDG2Used'] = df_EC['IsDG2Used'].apply(replace_value)
                df_EC['IsDG3Used'] = df_EC['IsDG3Used'].apply(replace_value)
                df_EC['IsDG4Used'] = df_EC['IsDG4Used'].apply(replace_value)
                df_EC['IsDG5Used'] = df_EC['IsDG5Used'].apply(replace_value)
                df_EC['IsDG6Used'] = df_EC['IsDG6Used'].apply(replace_value)

                log_Activity.append("value of IsDG1Used to IsDG6Used is relaced with 1 and 0")

                df_EC[['IsDG1Used','IsDG2Used', 'IsDG3Used', 'IsDG4Used', 'IsDG5Used', 'IsDG6Used']] = df_EC[['IsDG1Used','IsDG2Used', 'IsDG3Used', 'IsDG4Used', 'IsDG5Used', 'IsDG6Used']].apply(pd.to_numeric)
                log_Activity.append("IsDG1Used to IsDG6Used is converted to numeric") 
                
                # Apply the parameters
                df_filtered = df_EC[(df_EC['SpanName'] == 'Propelling') & (df_EC['MachineID'] == 6)]
                log_Activity.append("Parameters are applied in Event & Consumtion dataset")

                # temp_filtered_data_loc = os.path.join(obj_DataLakeUtility.temp_directory, "Filtered_Data.csv")
                # df_filtered.to_csv(temp_filtered_data_loc,index=False)
                # log_Activity.append('Filtered DGBoiler Data has been converted to CSV')

                df_Prev = pd.read_csv(Final_processed_file_path, header=0) #old proceesed file read
                log_Activity.append("Old Processed file read complete")

                result_df = pd.concat([df_Prev,df_filtered], ignore_index=True) #merge today's data in old processed file
                log_Activity.append("Today's data has been appnded in previous processed file")

                #-----------------------------------------------------------------------------------------------------
                #                          Storing the Data for Power Bi Dashboard Start                             |
                #-----------------------------------------------------------------------------------------------------

                temp_final_data_loc = os.path.join(obj_DataLakeUtility.temp_directory, "Final_Data.csv")
                result_df.to_csv(temp_final_data_loc,index=False)
                log_Activity.append('Final DGBoiler Data has been converted to CSV')

                Filtered_data_file_upload = obj_DataLakeUtility.upload_blob(destination_container_name, blob_path="ABLOG_FINAL/DG-Boiler", temp_file_abs_path=temp_final_data_loc, blob_name="DGBoiler_processed_data.csv")
                log_Activity.append("Filtered DGBoiler Data has been uploaded in Blob location")

                #---------------------------------------------------------------------------------------------------
                #                        Storing the Data for Power Bi Dashboard End                               |
                #---------------------------------------------------------------------------------------------------

                #---------------------------------------------------------------------------------------------------
                #                        Storing the data for Excel Dashboard Start                                |
                #---------------------------------------------------------------------------------------------------

                result_df2 = result_df[['VoyageNo', 'Division','IMONo','VesselName','EventDateTime', 'Event Date','EventName', 'SpanName','SpanTime2','SeawaterTemperature','machine_name','IsDG1Used','IsDG2Used', 'IsDG3Used', 'IsDG4Used', 'IsDG5Used', 'IsDG6Used','AveOutput','oil_name','OilConsumptionRate']]
                col_sum = ['IsDG1Used','IsDG2Used', 'IsDG3Used', 'IsDG4Used', 'IsDG5Used', 'IsDG6Used']
                result_df2['Dg Using'] = result_df2[col_sum].sum(axis=1)

                result_df2.rename(columns={'SpanTime2': 'Span Time'}, inplace=True)
                result_df2.rename(columns={'machine_name': 'Machine Name'}, inplace=True)
                result_df2.rename(columns={'oil_name': 'Oil Name'}, inplace=True)
                result_df2.drop(['IsDG1Used','IsDG2Used', 'IsDG3Used', 'IsDG4Used', 'IsDG5Used', 'IsDG6Used'], inplace=True, axis=1)

                result_df2 = result_df2[['VoyageNo','Division','IMONo', 'VesselName','EventDateTime', 'Event Date','EventName','SpanName','Span Time','SeawaterTemperature', 'Machine Name', 'Dg Using', 'AveOutput', 'Oil Name', 'OilConsumptionRate']]

                result_df2['EventDateTime'] = result_df2['EventDateTime'].astype('str')
                result_df2['Span Time'] = result_df2['Span Time'].astype('str')
                temp_final_excel_data_loc = os.path.join(obj_DataLakeUtility.temp_directory, "Final_Excel_Data.xlsx")
                result_df2.to_excel(temp_final_excel_data_loc,index=False)
                log_Activity.append('Final DGBoiler Data has been converted to Excel')

                Filtered_data_excel_file_upload = obj_DataLakeUtility.upload_blob(destination_container_name, blob_path="ABLOG_FINAL/DG-Boiler", temp_file_abs_path=temp_final_excel_data_loc, blob_name="Aux_Boiler_Consumtion_data.xlsx")
                log_Activity.append('Filtered Aux_Boiler excel Data has been uploaded in Blob location')

                #---------------------------------------------------------------------------------------------------
                #                        Storing the data for Excel Dashboard End                                |
                #---------------------------------------------------------------------------------------------------

                obj_DataLakeUtility.delete_dir()

                log_Activity.append('--------------------------------')
                log_Activity.append("Copy & Delete blob function is initializing...")
                
                obj_DataLakeUtility.copy_and_delete()

                log_Activity.append("Copy & Delete blob function is completed successfully")
                log_Activity.append('--------------------------------')
        except Exception as e:
                logging.exception(str(e))
                log_Activity.append(e)
        finally:
                Activity_Function_Log_file_Upload = obj_DataLakeUtility.upload_log(log_data_string=str(log_Activity),blob_name="Activity_Function_Log.log",blob_path=blob_path,destination_container_name=destination_container_name)

        return "Activity Completed"